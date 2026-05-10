"""One-shot: rewrite dummy invoice test requirements and dL/da/db targets.

Rules:
- Tests pulled only from lot_test_results.csv (no hallucinated method ids).
- Per color_family the allowed methods = methods present in inventory tests.
- If no inventory tests for the line's color -> no test (empty).
- Otherwise deterministic mix by invoice_line_id hash:
    bucket 0 (~20%) -> no test
    bucket 1-5 (~50%) -> single required test
    bucket 6-9 (~30%) -> super test (2 or more methods, joined by ',')
- dL/da/db targets are anchored to real inventory tests, then jittered by a
  small deterministic variance so invoice values are close but not exact lot
  copies. Blank-test invoice lines use method_i_a as the base-test anchor.
"""
from __future__ import annotations

import hashlib
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
import pandas as pd

BACKEND = Path(__file__).parent
STITCHED = BACKEND / "stitched_outputs"

TESTS_FILE = STITCHED / "lot_test_results.csv"
CSV_TARGETS = [
    STITCHED / "dummy_invoice_lines.csv",
    STITCHED / "dummy_invoice_lines.csv.baseline",
    STITCHED / "bulk_dummy_invoice_lines_240.csv",
    STITCHED / "bulk_dummy_invoice_lines_240.csv.baseline",
]
WORKBOOK_TARGETS = [
    STITCHED / "stitched_dataset.xlsx",
    STITCHED / "bulk_dummy_invoices_240.xlsx",
]
SMALL_HEADERS_FILE = STITCHED / "dummy_invoice_headers.csv"
SMALL_LINES_FILE = STITCHED / "dummy_invoice_lines.csv"
METHOD_ORDER = [
    "method_i_a",
    "method_i_b",
    "method_ii",
    "method_iii",
    "method_iv_a",
    "method_iv_b",
    "method_v_a",
    "method_v_b",
    "method_vi_a",
    "method_vi_b",
]
BASE_TEST_METHOD = "method_i_a"
TARGET_COLUMNS = ["target_method_id", "target_delta_l", "target_delta_a", "target_delta_b"]
DEMO_TARGET_SOURCE_ID = "BULK-INV-5012-L02"
DEMO_TARGET_COPY_IDS = ["BULK-INV-5014-L03", "BULK-INV-5010-L02"]


def clean(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    return str(value).strip()


def color_key(value: Any) -> str:
    return clean(value).upper()


def method_sort_key(method_id: str) -> tuple[int, str]:
    try:
        return (METHOD_ORDER.index(method_id), method_id)
    except ValueError:
        return (len(METHOD_ORDER), method_id)


def build_inventory_maps() -> tuple[
    dict[str, list[str]],
    dict[tuple[str, str], list[dict[str, Any]]],
]:
    tests = pd.read_csv(TESTS_FILE)
    methods_by_color: dict[str, list[str]] = {}
    tests_by_color_method: dict[tuple[str, str], list[dict[str, Any]]] = {}
    for color, grp in tests.groupby(tests["color_family"].apply(color_key)):
        methods = sorted(
            {clean(m) for m in grp["method_id"].dropna().tolist() if clean(m)},
            key=method_sort_key,
        )
        if methods:
            methods_by_color[color] = methods
        for method_id, mgrp in grp.groupby(grp["method_id"].apply(clean)):
            if not method_id:
                continue
            rows = []
            for _, row in mgrp.iterrows():
                if pd.isna(row.get("delta_l")) or pd.isna(row.get("delta_a")) or pd.isna(row.get("delta_b")):
                    continue
                rows.append(
                    {
                        "lot_no": clean(row.get("lot_no")),
                        "delta_l": float(row.get("delta_l")),
                        "delta_a": float(row.get("delta_a")),
                        "delta_b": float(row.get("delta_b")),
                    }
                )
            if rows:
                tests_by_color_method[(color, method_id)] = rows
    return methods_by_color, tests_by_color_method


def bucket(line_id: str) -> int:
    h = hashlib.md5(str(line_id).encode("utf-8")).hexdigest()
    return int(h[:2], 16) % 10  # 0..9


def pick_methods(line_id: str, available: list[str]) -> str:
    if not available:
        return ""
    b = bucket(line_id)
    if b == 0:
        return ""
    if b <= 5:
        idx = int(hashlib.md5(f"{line_id}:single".encode()).hexdigest()[:8], 16) % len(available)
        return available[idx]
    # super: 2 or more
    h = hashlib.md5(f"{line_id}:super".encode()).hexdigest()
    n_max = min(len(available), 3)
    if n_max < 2:
        return available[0]
    n = 2 + (int(h[:2], 16) % (n_max - 1))  # 2..n_max
    seed = int(h, 16)
    pool = list(available)
    chosen: list[str] = []
    for _ in range(n):
        i = seed % len(pool)
        chosen.append(pool.pop(i))
        seed //= max(len(pool) + 1, 1)
    # preserve canonical order
    order = {m: i for i, m in enumerate(available)}
    chosen.sort(key=lambda m: order[m])
    return ",".join(chosen)


def parse_methods(value: str, available: list[str]) -> list[str]:
    allowed = set(available)
    out: list[str] = []
    for chunk in str(value or "").replace(";", ",").replace("|", ",").split(","):
        method_id = chunk.strip()
        if method_id in allowed and method_id not in out:
            out.append(method_id)
    return out


def selected_match_method(method_value: str, available: list[str]) -> str:
    required = parse_methods(method_value, available)
    if required:
        return sorted(required, key=lambda m: (-(METHOD_ORDER.index(m) if m in METHOD_ORDER else -1), required.index(m)))[0]
    if BASE_TEST_METHOD in available:
        return BASE_TEST_METHOD
    return available[0] if available else ""


def pick_anchor_test(line_id: str, method_id: str, rows: list[dict[str, Any]]) -> dict[str, Any]:
    seed = int(hashlib.md5(f"{line_id}:{method_id}:anchor".encode("utf-8")).hexdigest(), 16)
    return rows[seed % len(rows)]


def variance(line_id: str, axis: str) -> float:
    digest = hashlib.md5(f"{line_id}:{axis}:variance".encode("utf-8")).hexdigest()
    sign = -1 if int(digest[0], 16) % 2 else 1
    magnitude = 0.05 + (int(digest[1:5], 16) / 0xFFFF) * 0.19
    return sign * magnitude


def jittered_delta(
    line_id: str,
    color: str,
    method_id: str,
    anchor: dict[str, Any],
    tests_by_color_method: dict[tuple[str, str], list[dict[str, Any]]],
) -> tuple[float, float, float]:
    dl = round(float(anchor["delta_l"]) + variance(line_id, "dL"), 2)
    da = round(float(anchor["delta_a"]) + variance(line_id, "dA"), 2)
    db = round(float(anchor["delta_b"]) + variance(line_id, "dB"), 2)

    lot_triples = {
        (round(float(r["delta_l"]), 2), round(float(r["delta_a"]), 2), round(float(r["delta_b"]), 2))
        for r in tests_by_color_method.get((color, method_id), [])
    }
    if (dl, da, db) in lot_triples:
        bump = 0.07 if variance(line_id, "avoid-exact") >= 0 else -0.07
        db = round(db + bump, 2)
    return dl, da, db


def apply_bulk_demo_overrides(df: pd.DataFrame, updates: dict[str, dict[str, Any]]) -> None:
    """Pin yellow shortage lines to the same strict target signature for repeatable demo outcomes."""
    ids = set([DEMO_TARGET_SOURCE_ID, *DEMO_TARGET_COPY_IDS])
    if "invoice_line_id" not in df.columns or not ids.issubset(set(df["invoice_line_id"].astype(str))):
        return

    src_row = df.loc[df["invoice_line_id"].astype(str) == DEMO_TARGET_SOURCE_ID, TARGET_COLUMNS].iloc[0]
    src_update = {col: src_row[col] for col in TARGET_COLUMNS}

    for target_id in DEMO_TARGET_COPY_IDS:
        mask = df["invoice_line_id"].astype(str) == target_id
        for col in TARGET_COLUMNS:
            df.loc[mask, col] = src_update[col]
        updates[target_id] = dict(src_update)


def rewrite_csv(
    path: Path,
    methods_by_color: dict[str, list[str]],
    tests_by_color_method: dict[tuple[str, str], list[dict[str, Any]]],
) -> tuple[dict[str, int], dict[str, dict[str, Any]]]:
    df = pd.read_csv(path)
    counts = {"none": 0, "single": 0, "super": 0, "no_inventory": 0}
    updates: dict[str, dict[str, Any]] = {}
    for _, row in df.iterrows():
        color = color_key(row.get("color_family", ""))
        line_id = clean(row.get("invoice_line_id", ""))
        available = methods_by_color.get(color, [])
        val = pick_methods(line_id, available)
        if not available:
            counts["no_inventory"] += 1
            updates[line_id] = {"target_method_id": ""}
        elif not val:
            counts["none"] += 1
            method_id = selected_match_method("", available)
            anchor = pick_anchor_test(line_id, method_id, tests_by_color_method[(color, method_id)])
            dl, da, db = jittered_delta(line_id, color, method_id, anchor, tests_by_color_method)
            updates[line_id] = {
                "target_method_id": "",
                "target_delta_l": dl,
                "target_delta_a": da,
                "target_delta_b": db,
            }
        elif "," in val:
            counts["super"] += 1
            method_id = selected_match_method(val, available)
            anchor = pick_anchor_test(line_id, method_id, tests_by_color_method[(color, method_id)])
            dl, da, db = jittered_delta(line_id, color, method_id, anchor, tests_by_color_method)
            updates[line_id] = {
                "target_method_id": val,
                "target_delta_l": dl,
                "target_delta_a": da,
                "target_delta_b": db,
            }
        else:
            counts["single"] += 1
            method_id = selected_match_method(val, available)
            anchor = pick_anchor_test(line_id, method_id, tests_by_color_method[(color, method_id)])
            dl, da, db = jittered_delta(line_id, color, method_id, anchor, tests_by_color_method)
            updates[line_id] = {
                "target_method_id": val,
                "target_delta_l": dl,
                "target_delta_a": da,
                "target_delta_b": db,
            }
    for col in TARGET_COLUMNS:
        if col not in df.columns:
            df[col] = ""
    for idx, row in df.iterrows():
        line_id = clean(row.get("invoice_line_id", ""))
        update = updates.get(line_id, {})
        for col, val in update.items():
            df.at[idx, col] = val
    if "bulk_dummy_invoice_lines_240" in path.name:
        apply_bulk_demo_overrides(df, updates)
    df.to_csv(path, index=False)
    return counts, updates


def update_workbook(path: Path, updates: dict[str, dict[str, Any]]) -> int:
    wb = load_workbook(path)
    if "invoice_lines" not in wb.sheetnames:
        return 0
    ws = wb["invoice_lines"]
    headers = {clean(cell.value): idx + 1 for idx, cell in enumerate(ws[1])}
    if "invoice_line_id" not in headers:
        return 0
    for col in TARGET_COLUMNS:
        if col not in headers:
            ws.cell(row=1, column=ws.max_column + 1, value=col)
            headers[col] = ws.max_column
    changed = 0
    for row_idx in range(2, ws.max_row + 1):
        line_id = clean(ws.cell(row=row_idx, column=headers["invoice_line_id"]).value)
        update = updates.get(line_id)
        if not update:
            continue
        for col, val in update.items():
            ws.cell(row=row_idx, column=headers[col], value=val)
        changed += 1
    wb.save(path)
    return changed


def md_cell(value: Any) -> str:
    value_text = clean(value)
    return value_text.replace("|", "\\|") if value_text else ""


def update_markdown_invoices() -> int:
    if not SMALL_HEADERS_FILE.exists() or not SMALL_LINES_FILE.exists():
        return 0
    headers = pd.read_csv(SMALL_HEADERS_FILE)
    lines = pd.read_csv(SMALL_LINES_FILE)
    changed = 0
    for _, header in headers.iterrows():
        invoice_id = clean(header.get("invoice_id"))
        invoice_number = clean(header.get("invoice_number")) or invoice_id.replace("INV-", "")
        if not invoice_id and not invoice_number:
            continue
        scoped = lines[lines["invoice_id"].apply(clean) == invoice_id].copy()
        if scoped.empty:
            continue
        scoped = scoped.sort_values("line_no")
        out = [
            f"# Invoice {invoice_number}",
            "",
            f"- Customer: `{md_cell(header.get('customer_name'))}`",
            f"- Date: `{md_cell(header.get('invoice_date'))}`",
            f"- Kind: `{md_cell(header.get('invoice_kind'))}`",
            "",
            "| Line | Product | Grade | Std | App | Test | dL | da | db | Qty kg | Match |",
            "| --- | --- | --- | --- | --- | --- | ---: | ---: | ---: | ---: | --- |",
        ]
        for _, line in scoped.iterrows():
            out.append(
                "| "
                + " | ".join(
                    [
                        md_cell(line.get("line_no")),
                        md_cell(line.get("product_description")),
                        md_cell(line.get("grade")),
                        md_cell(line.get("standard_code")),
                        md_cell(line.get("application")),
                        md_cell(line.get("target_method_id")),
                        md_cell(line.get("target_delta_l")),
                        md_cell(line.get("target_delta_a")),
                        md_cell(line.get("target_delta_b")),
                        md_cell(line.get("qty_kg")),
                        md_cell(line.get("inventory_match_status")),
                    ]
                )
                + " |"
            )
        (STITCHED / f"INV-{invoice_number}.md").write_text("\n".join(out) + "\n", encoding="utf-8")
        changed += 1
    return changed


def main() -> None:
    methods_by_color, tests_by_color_method = build_inventory_maps()
    print("Allowed methods per color:")
    for k, v in sorted(methods_by_color.items()):
        print(f"  {k}: {v}")
    print()
    all_updates: dict[str, dict[str, Any]] = {}
    for tgt in CSV_TARGETS:
        c, updates = rewrite_csv(tgt, methods_by_color, tests_by_color_method)
        all_updates.update(updates)
        print(f"{tgt.name}: {c}")
    for tgt in WORKBOOK_TARGETS:
        changed = update_workbook(tgt, all_updates)
        print(f"{tgt.name}: updated {changed} invoice_lines rows")
    changed = update_markdown_invoices()
    print(f"invoice markdowns: updated {changed} files")


if __name__ == "__main__":
    main()
